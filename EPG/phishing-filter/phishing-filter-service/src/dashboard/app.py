import streamlit as st
import sys
from pathlib import Path
import email
from email.message import Message

# Add project root to python path
sys.path.append(str(Path(__file__).resolve().parent.parent.parent))

from src.utils.email_parser import EmailParser
from src.engines.header_analyzer import HeaderAnalyzer
from src.engines.url_analyzer import URLAnalyzer
from src.engines.nlp_analyzer import NLPAnalyzer
from src.engines.aggregator import HybridRiskAggregator
from src.config import LEGITIMATE_BRANDS, DOMAIN_EXCEL

# Set page configuration with high-tech SOC style
st.set_page_config(
    page_title="SOC Phishing Analyzer Platform",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom Premium Styling
st.markdown("""
<style>
    /* Dark Slate Premium Background styling */
    .reportview-container {
        background: #0e1117;
    }
    
    /* Metrics panel card design */
    .threat-card {
        padding: 1.5rem;
        border-radius: 12px;
        background-color: #1f2937;
        border: 1px solid #374151;
        margin-bottom: 1rem;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
    }
    
    /* Dynamic color highlights */
    .risk-clean {
        border-left: 6px solid #10b981;
    }
    .risk-suspicious {
        border-left: 6px solid #f59e0b;
    }
    .risk-phishing {
        border-left: 6px solid #ef4444;
    }
    
    .threat-header {
        font-size: 1.25rem;
        font-weight: 600;
        margin-bottom: 0.75rem;
        color: #f3f4f6;
    }
    
    .threat-value {
        font-size: 2.25rem;
        font-weight: 700;
        line-height: 1;
        color: #ffffff;
    }
    
    .threat-subtitle {
        font-size: 0.875rem;
        color: #9ca3af;
        margin-top: 0.25rem;
    }
    
    .reasons-list {
        background-color: #111827;
        border-radius: 8px;
        padding: 1rem;
        border-left: 4px solid #f59e0b;
        color: #e5e7eb;
    }
</style>
""", unsafe_allow_html=True)

# Cache analyzers and models loading
@st.cache_resource
def init_detection_pipeline():
    brands = list(LEGITIMATE_BRANDS)
    if DOMAIN_EXCEL.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(DOMAIN_EXCEL, read_only=True)
            sheet = wb.active
            extra_domains = []
            for row in sheet.iter_rows(values_only=True):
                if row and row[0]:
                    val = str(row[0]).strip().lower()
                    if val and val != "domain" and "." in val:
                        extra_domains.append(val)
            if extra_domains:
                brands.extend(extra_domains)
                brands = list(set(brands))
        except Exception:
            pass
            
    header_analyzer = HeaderAnalyzer(brand_list=brands)
    url_analyzer = URLAnalyzer(brand_list=brands)
    nlp_analyzer = NLPAnalyzer()
    aggregator = HybridRiskAggregator()
    
    return header_analyzer, url_analyzer, nlp_analyzer, aggregator

header_analyzer, url_analyzer, nlp_analyzer, aggregator = init_detection_pipeline()

# Title and subtitle
st.markdown("<h1>🛡️ SOC Advanced Phishing Detection Pipeline</h1>", unsafe_allow_html=True)
st.markdown("<p style='color:#9ca3af; font-size:1.15rem; margin-top:-0.5rem;'>Enterprise-Grade Dual-Engine Secure Email Gateway (SEG) Simulation Platform</p>", unsafe_allow_html=True)
st.write("---")

# Sidebar Configuration and Preset Attacks
st.sidebar.image("https://img.icons8.com/nolan/96/shield.png", width=80)
st.sidebar.markdown("### ⚙️ Pipeline Controls")

# Trigger reload models manually
if st.sidebar.button("🔄 Reload ML Models"):
    nlp_analyzer.load_models()
    aggregator.load_model()
    st.sidebar.success("Models reloaded!")

st.sidebar.markdown("### 🧪 Select Threat Attack Preset")
presets = {
    "None": None,
    "1. Display Name Spoofing (PayPal)": {
        "sender": '"PayPal Security Alert" <alerts-security-system@phish-login-portal.net>',
        "receiver": "saif@enterprise.com",
        "subject": "Immediate Action Required: Your Account has been Restricted",
        "body_text": "We detected unauthorized login attempts to your account. Your account has been temporarily restricted. Please login within 24 hours to verify your identity, otherwise your account will be permanently suspended.",
        "body_html": 'We detected unauthorized login attempts. Please <a href="http://login-paypal-verify-identity-login.com/security/signin.php">login to PayPal to verify your account</a> immediately.',
        "reply_to": "support-paypal@gmail.com",
        "x_mailer": "PHP mailer script v5.4",
        "auth_headers": {"spf": "FAIL", "dkim": "FAIL", "dmarc": "FAIL", "raw_auth_results": "spf=fail; dmarc=fail"}
    },
    "2. BEC CEO Wire Impersonation": {
        "sender": '"John Doe (CEO)" <ceo-john-doe@gmail.com>',
        "receiver": "finance-dept@enterprise.com",
        "subject": "Confidential Request - Wire Transfer Required Immediately",
        "body_text": "Are you at your desk? I am in a confidential meeting right now and need you to process an urgent wire transfer of $45,000 for an ongoing acquisition. Please reply immediately so I can provide the bank details. Keep this confidential.",
        "body_html": "Are you at your desk? I am in a confidential meeting right now and need you to process an urgent wire transfer of $45,000 for an ongoing acquisition. Please reply immediately so I can provide the bank details. Keep this confidential.",
        "reply_to": "ceo-office-direct-reply@hackermail.cc",
        "x_mailer": "Outlook for iOS",
        "auth_headers": {"spf": "PASS", "dkim": "NONE", "dmarc": "NONE", "raw_auth_results": "spf=pass"}
    },
    "3. Homograph Unicode Attack (Cyrillic)": {
        "sender": '"Microsoft Support" <support@rnicrosoft.com>',
        "receiver": "saif@enterprise.com",
        "subject": "Security Alert: Microsoft Account Password Expiry Warning",
        "body_text": "Your password is set to expire in 48 hours. Please keep your current password by visiting the service verification portal.",
        "body_html": 'Please secure your account at <a href="http://xn--mcrosoft-12a.com/renew">microsoft.com service portal</a> (Note: raw unicode: microsoft.com using Cyrillic i)',
        "reply_to": "",
        "x_mailer": "Microsoft Exchange SMTP",
        "auth_headers": {"spf": "PASS", "dkim": "PASS", "dmarc": "PASS", "raw_auth_results": "spf=pass; dkim=pass"}
    },
    "4. Clean Legitimate Newsletter": {
        "sender": '"GitHub Community" <noreply@github.com>',
        "receiver": "saif@enterprise.com",
        "subject": "GitHub Changelog: New features in code scanning and security keys",
        "body_text": "Hello, here are the updates on GitHub for May 2026. Code scanning is now faster, and we support multi-factor security keys for ssh logins. Read the details on our blog.",
        "body_html": 'Hello, read the details on <a href="https://github.blog/changelog">GitHub Blog</a>.',
        "reply_to": "noreply@github.com",
        "x_mailer": "SendGrid-Mailer",
        "auth_headers": {"spf": "PASS", "dkim": "PASS", "dmarc": "PASS", "raw_auth_results": "spf=pass; dkim=pass; dmarc=pass"}
    }
}

preset_name = st.sidebar.selectbox("Choose a sample preset:", list(presets.keys()))
preset_data = presets[preset_name]

# Main Workspace Layout: Two Panels
col_input, col_results = st.columns([1, 1])

with col_input:
    st.markdown("### 📥 Email Input Gateway")
    
    # Input tabs
    tab_eml, tab_text = st.tabs(["📁 Upload .eml File", "📝 Paste Raw Content / Metadata"])
    
    email_data = None
    
    with tab_eml:
        uploaded_file = st.file_uploader("Upload an RFC-822 standard email file (.eml)", type=["eml", "txt"])
        if uploaded_file is not None:
            try:
                raw_bytes = uploaded_file.read()
                raw_str = raw_bytes.decode('utf-8', errors='ignore')
                email_data = EmailParser.parse_from_string(raw_str)
                st.success(f"Successfully parsed email file: {uploaded_file.name}")
            except Exception as e:
                st.error(f"Error parsing EML file: {e}")

    with tab_text:
        # If preset is selected, populate these fields
        p_sender = preset_data["sender"] if preset_data else ""
        p_subject = preset_data["subject"] if preset_data else ""
        p_body = preset_data["body_html"] if preset_data else ""
        p_reply = preset_data["reply_to"] if preset_data else ""
        p_mailer = preset_data["x_mailer"] if preset_data else ""
        p_spf = preset_data["auth_headers"]["spf"] if preset_data else "PASS"
        p_dkim = preset_data["auth_headers"]["dkim"] if preset_data else "PASS"
        p_dmarc = preset_data["auth_headers"]["dmarc"] if preset_data else "PASS"

        with st.form("manual_input_form"):
            sender = st.text_input("From Header:", value=p_sender, placeholder="e.g. PayPal Security <support@paypal.com>")
            subject = st.text_input("Subject Line:", value=p_subject, placeholder="e.g. Action Required: Account Locked")
            
            col_nested1, col_nested2 = st.columns(2)
            with col_nested1:
                reply_to = st.text_input("Reply-To Header:", value=p_reply)
                x_mailer_input = st.text_input("X-Mailer / User-Agent:", value=p_mailer)
            with col_nested2:
                spf = st.selectbox("SPF Result:", ["PASS", "FAIL", "NEUTRAL", "NONE"], index=["PASS", "FAIL", "NEUTRAL", "NONE"].index(p_spf))
                dkim = st.selectbox("DKIM Result:", ["PASS", "FAIL", "NONE"], index=["PASS", "FAIL", "NONE"].index(p_dkim))
                dmarc = st.selectbox("DMARC Result:", ["PASS", "FAIL", "NONE"], index=["PASS", "FAIL", "NONE"].index(p_dmarc))

            body_html = st.text_area("Email Body (HTML or Plain Text):", value=p_body, height=180)
            
            submitted = st.form_submit_button("🛡️ Scan Input Payload")
            if submitted or (preset_data is not None and not uploaded_file):
                # Build mock data
                raw_urls = []
                # Simple extraction of hrefs from user html input
                from bs4 import BeautifulSoup
                soup = BeautifulSoup(body_html, "html.parser")
                for a in soup.find_all("a", href=True):
                    raw_urls.append({"href": a["href"], "text": a.get_text()})
                
                # Check for regex urls
                import re
                url_pattern = re.compile(r'https?://[^\s<>"]+|www\.[^\s<>"]+')
                found_urls = url_pattern.findall(body_html)
                existing_hrefs = {u["href"] for u in raw_urls}
                for u in found_urls:
                    cleaned_u = u.rstrip('.,;:)("]')
                    if cleaned_u not in existing_hrefs:
                        raw_urls.append({"href": cleaned_u, "text": cleaned_u})
                        existing_hrefs.add(cleaned_u)

                email_data = {
                    "sender": sender,
                    "receiver": "user@internal-domain.com",
                    "date": "Wed, 20 May 2026 21:00:00 +0300",
                    "subject": subject,
                    "reply_to": reply_to,
                    "x_mailer": x_mailer_input,
                    "body_text": soup.get_text(),
                    "body_html": body_html,
                    "urls": raw_urls,
                    "auth_headers": {
                        "spf": spf,
                        "dkim": dkim,
                        "dmarc": dmarc,
                        "raw_auth_results": f"spf={spf.lower()}; dkim={dkim.lower()}; dmarc={dmarc.lower()}"
                    }
                }

# Results Column Display
with col_results:
    st.markdown("### 📊 Live Pipeline Verdict Reports")
    
    if email_data is not None:
        # Run analyzers
        h_res = header_analyzer.analyze(email_data)
        u_res = url_analyzer.analyze(email_data)
        n_res = nlp_analyzer.analyze(email_data)
        
        # Run aggregator
        agg_res = aggregator.aggregate(h_res, u_res, n_res)
        
        verdict = agg_res["verdict"]
        risk_score = agg_res["risk_score"]
        confidence = agg_res["confidence"]
        reasons = agg_res["reasons"]
        mode = agg_res["mode"]
        
        # Color coding class — binary only: PHISHING (red) or CLEAN (green)
        card_class = "risk-clean"
        verdict_color = "#10b981"
        if verdict == "PHISHING":
            card_class = "risk-phishing"
            verdict_color = "#ef4444"
            
        # Verdict Dashboard Card
        st.markdown(f"""
        <div class="threat-card {card_class}">
            <div class="threat-header">PIPELINE VERDICT</div>
            <div class="threat-value" style="color: {verdict_color};">{verdict}</div>
            <div class="threat-subtitle">Aggregator Verdict engine powered by: <b>{mode}</b></div>
        </div>
        """, unsafe_allow_html=True)
        
        # Metrics breakdown row
        col_m1, col_m2, col_m3 = st.columns(3)
        with col_m1:
            st.metric("Unified Threat Score", f"{risk_score:.1f}%", help="Aggregation of heuristics and ML models")
        with col_m2:
            st.metric("Aggregator Confidence", f"{confidence * 100:.0f}%", help="Aggregator model confidence level")
        with col_m3:
            st.metric("Detected Links Count", f"{len(email_data['urls'])} links")
            
        # Explanations block
        st.markdown("#### 💡 Explainable AI Indicators (XAI)")
        if reasons:
            reasons_html = "".join([f"<li>🔴 {r}</li>" for r in reasons])
            st.markdown(f"""
            <div class="reasons-list">
                <ul style="margin: 0; padding-left: 1.25rem;">
                    {reasons_html}
                </ul>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class="reasons-list" style="border-left: 4px solid #10b981;">
                ✅ No suspicious header, URL, or NLP intent indicators triggered.
            </div>
            """, unsafe_allow_html=True)

        st.write("---")
        
        # Engine Breakdown Tabs
        tab_nlp, tab_url, tab_header = st.tabs([
            "🧠 NLP Intent", 
            "🔗 URL Obfuscation", 
            "📬 Header Integrity"
        ])
        
        with tab_nlp:
            st.markdown("#### NLP Semantic Intent Classifier Metrics")
            nlp_scores = {
                "Phishing Intent": n_res["phishing_probability"],
                "Urgency / Coercion": n_res["urgency_score"],
                "Credential Theft Request": n_res["credential_theft_score"],
                "BEC / Invoice Scam": n_res["BEC_probability"]
            }
            
            # Use native streamlit bar chart
            st.bar_chart(nlp_scores, horizontal=True)
            
            # Print NLP specific features
            st.json({
                "nlp_mode": n_res.get("mode", "Unknown"),
                "features_extracted": n_res.get("features", {})
            })

        with tab_url:
            st.markdown("#### Static Link Analysis Reports")
            urls_checked = u_res.get("urls_checked", [])
            
            if not urls_checked:
                st.info("No links found in the email payload.")
            else:
                for idx, url_info in enumerate(urls_checked):
                    score = url_info['score']
                    
                    with st.expander(f"🌐 Link #{idx+1}: {url_info['domain']} (Risk Score: {score:.1f})"):
                        st.markdown(f"**URL:** `{url_info['url']}`")
                        st.markdown(f"**Anchor Display Text:** `\"{url_info['text']}\"`")
                        st.markdown(f"**Domain:** `{url_info['domain']}` | **TLD:** `.{url_info['tld']}`")
                        st.markdown(f"**Shannon Domain Entropy:** `{url_info['entropy']:.3f}`")
                        
                        if url_info['reasons']:
                            st.write("Flags Raised:")
                            for r in url_info['reasons']:
                                st.markdown(f"- ⚠️ <span style='color:red;'>{r}</span>", unsafe_allow_html=True)
                        else:
                            st.success("No URL flags triggered.")

        with tab_header:
            st.markdown("#### Header Verification & Alignment")
            details = h_res.get("details", {})
            
            col_h1, col_h2 = st.columns(2)
            with col_h1:
                st.markdown(f"**Display Name:** `{details.get('display_name')}`")
                st.markdown(f"**Actual Envelope Sender:** `{details.get('sender_address')}`")
                st.markdown(f"**Reply-To Domain:** `{details.get('reply_to_domain') or 'None'}`")
                st.markdown(f"**User Agent / X-Mailer:** `{details.get('x_mailer') or 'None'}`")
            with col_h2:
                # Security indicators
                spf_val = details.get("auth_headers", {}).get("spf", "NONE")
                dkim_val = details.get("auth_headers", {}).get("dkim", "NONE")
                dmarc_val = details.get("auth_headers", {}).get("dmarc", "NONE")
                
                def draw_badge(name, status):
                    color = "green" if status == "PASS" else "red" if status == "FAIL" else "orange" if status == "NEUTRAL" else "gray"
                    return f"<span style='background-color:{color}; color:white; padding:2px 8px; border-radius:4px; font-weight:bold; font-size:0.8rem;'>{status}</span>"
                
                st.markdown(f"**SPF Alignment:** {draw_badge('SPF', spf_val)}", unsafe_allow_html=True)
                st.markdown(f"**DKIM Validation:** {draw_badge('DKIM', dkim_val)}", unsafe_allow_html=True)
                st.markdown(f"**DMARC Alignment:** {draw_badge('DMARC', dmarc_val)}", unsafe_allow_html=True)
                
            st.markdown("---")
            st.write("Raw Heuristics features:")
            st.json(h_res.get("features", {}))
            
    else:
        st.info("👈 Upload an email file or choose a Threat Attack Preset from the sidebar to inspect the pipeline results.")
