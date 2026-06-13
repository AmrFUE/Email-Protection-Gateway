from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks
from pydantic import BaseModel, Field
from typing import List, Dict, Any, Optional
import uvicorn
import sys
from pathlib import Path

# Add project root to path
sys.path.append(str(Path(__file__).resolve().parent.parent))

from src.utils.email_parser import EmailParser
from src.engines.header_analyzer import HeaderAnalyzer
from src.engines.url_analyzer import URLAnalyzer
from src.engines.nlp_analyzer import NLPAnalyzer
from src.engines.aggregator import HybridRiskAggregator
from src.config import LEGITIMATE_BRANDS, DOMAIN_EXCEL

app = FastAPI(
    title="Advanced Phishing Detection Microservice",
    description="Enterprise-grade email scanning API using security heuristics and machine learning.",
    version="1.0.0"
)

# Initialize engines
# Load extra domains if Excel exists
brands = list(LEGITIMATE_BRANDS)
if DOMAIN_EXCEL.exists():
    try:
        from openpyxl import load_workbook
        wb = load_workbook(DOMAIN_EXCEL, read_only=True)
        sheet = wb.active
        extra_domains = []
        for row in sheet.iter_rows(values_only=True):
            if row and row[0]:
                val = str(row[0]).strip().lower()
                if val and val != "domain" and "." in val:
                    extra_domains.append(val)
        if extra_domains:
            brands.extend(extra_domains)
            brands = list(set(brands))
    except Exception:
        pass

header_analyzer = HeaderAnalyzer(brand_list=brands)
url_analyzer = URLAnalyzer(brand_list=brands)
nlp_analyzer = NLPAnalyzer()
aggregator = HybridRiskAggregator()

# Pydantic Schemas for API requests
class AuthHeaders(BaseModel):
    spf: str = Field(default="NONE", description="SPF check result (PASS/FAIL/NEUTRAL/NONE)")
    dkim: str = Field(default="NONE", description="DKIM check result (PASS/FAIL/NONE)")
    dmarc: str = Field(default="NONE", description="DMARC check result (PASS/FAIL/NONE)")
    raw_auth_results: str = Field(default="", description="Raw Authentication-Results or Received-SPF header content")

class UrlItem(BaseModel):
    href: str = Field(..., description="The link destination URL")
    text: str = Field(default="", description="The clickable display text of the link")

class EmailPayload(BaseModel):
    sender: str = Field(..., description="From header value (e.g. 'PayPal Support <hacker@evil.com>')")
    receiver: str = Field(default="", description="To header value")
    date: str = Field(default="", description="Date header value")
    subject: str = Field(default="", description="Subject line")
    reply_to: str = Field(default="", description="Reply-To header value")
    x_mailer: str = Field(default="", description="X-Mailer or User-Agent header value")
    body_text: str = Field(default="", description="Plain text body content")
    body_html: str = Field(default="", description="HTML body content")
    urls: List[UrlItem] = Field(default=[], description="Extracted URLs")
    auth_headers: AuthHeaders = Field(default_factory=AuthHeaders)

# Response schemas
class MetricDetails(BaseModel):
    header_score: float
    url_score: float
    nlp_score: float

class AnalysisResponse(BaseModel):
    verdict: str = Field(..., description="Verdicts: CLEAN, SUSPICIOUS, or PHISHING")
    risk_score: float = Field(..., description="Unified risk score from 0.0 to 100.0")
    confidence: float = Field(..., description="Confidence score from 0.50 to 1.00")
    reasons: List[str] = Field(..., description="Explainable AI indicators and reasons for the verdict")
    mode: str = Field(..., description="Execution mode (Heuristics or ML Ensemble)")
    metrics: MetricDetails
    engines: Dict[str, Any] = Field(..., description="Breakdown reports from each individual engine")

@app.on_event("startup")
def startup_event():
    # Reload models on startup to ensure we pick up fresh training weights
    nlp_analyzer.load_models()
    aggregator.load_model()

@app.get("/")
def read_root():
    return {
        "status": "active",
        "service": "Advanced Phishing Detection Pipeline",
        "endpoints": {
            "/analyze": "POST email JSON payload for full analysis",
            "/analyze-eml": "POST upload raw .eml file for full analysis",
            "/health": "GET health check"
        }
    }

@app.get("/health")
def health_check():
    return {
        "status": "healthy",
        "models_loaded": {
            "nlp_classifiers": nlp_analyzer.phishing_model is not None,
            "aggregator": aggregator.model is not None
        }
    }

@app.post("/analyze", response_model=AnalysisResponse)
def analyze_email_json(payload: EmailPayload):
    """Analyzes structured email data provided in JSON format."""
    email_data = payload.dict()
    
    # Run pipeline
    h_res = header_analyzer.analyze(email_data)
    u_res = url_analyzer.analyze(email_data)
    n_res = nlp_analyzer.analyze(email_data)
    
    agg_res = aggregator.aggregate(h_res, u_res, n_res)
    
    return {
        "verdict": agg_res["verdict"],
        "risk_score": agg_res["risk_score"],
        "confidence": agg_res["confidence"],
        "reasons": agg_res["reasons"],
        "mode": agg_res["mode"],
        "metrics": {
            "header_score": agg_res["metrics"]["header_score"],
            "url_score": agg_res["metrics"]["url_score"],
            "nlp_score": agg_res["metrics"]["nlp_score"]
        },
        "engines": {
            "header_analysis": h_res,
            "url_analysis": u_res,
            "nlp_analysis": n_res
        }
    }

@app.post("/analyze-eml", response_model=AnalysisResponse)
async def analyze_email_file(file: UploadFile = File(...)):
    """Accepts raw .eml files, parses them on-the-fly, and runs full analysis."""
    if not file.filename.lower().endswith(('.eml', '.txt', '.msg')):
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload a standard email (.eml, .txt) file.")
        
    try:
        contents = await file.read()
        raw_email_str = contents.decode('utf-8', errors='ignore')
        
        # Parse using EmailParser
        email_data = EmailParser.parse_from_string(raw_email_str)
        
        # Run pipeline
        h_res = header_analyzer.analyze(email_data)
        u_res = url_analyzer.analyze(email_data)
        n_res = nlp_analyzer.analyze(email_data)
        
        agg_res = aggregator.aggregate(h_res, u_res, n_res)
        
        return {
            "verdict": agg_res["verdict"],
            "risk_score": agg_res["risk_score"],
            "confidence": agg_res["confidence"],
            "reasons": agg_res["reasons"],
            "mode": agg_res["mode"],
            "metrics": {
                "header_score": agg_res["metrics"]["header_score"],
                "url_score": agg_res["metrics"]["url_score"],
                "nlp_score": agg_res["metrics"]["nlp_score"]
            },
            "engines": {
                "header_analysis": h_res,
                "url_analysis": u_res,
                "nlp_analysis": n_res
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse and analyze email file: {str(e)}")

if __name__ == "__main__":
    uvicorn.run("src.api.app:app", host="127.0.0.1", port=8000, reload=True)
