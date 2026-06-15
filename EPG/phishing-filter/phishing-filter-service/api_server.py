from fastapi import FastAPI, UploadFile, File, HTTPException
import uvicorn
import logging
import io

# Ensure root logger is at INFO so our capture handler sees all messages
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')

def _make_log_capture():
    buf = io.StringIO()
    handler = logging.StreamHandler(buf)
    handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s'))
    handler.setLevel(logging.DEBUG)
    return buf, handler

import tempfile
import os
from contextlib import asynccontextmanager
from typing import Dict, Any

from src.utils.email_parser import EmailParser
from src.engines.header_analyzer import HeaderAnalyzer
from src.engines.url_analyzer import URLAnalyzer
from src.engines.nlp_analyzer import NLPAnalyzer
from src.engines.aggregator import HybridRiskAggregator
from src.config import LEGITIMATE_BRANDS, DOMAIN_EXCEL

logger = logging.getLogger("phishing-filter")

# Global analyzer/aggregator instances
header_analyzer = None
url_analyzer = None
nlp_analyzer = None
aggregator = None

@asynccontextmanager
async def lifespan(app: FastAPI):
    global header_analyzer, url_analyzer, nlp_analyzer, aggregator
    # Initialize and load brand domain lists
    brands = list(LEGITIMATE_BRANDS)
    if DOMAIN_EXCEL.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(DOMAIN_EXCEL, read_only=True)
            sheet = wb.active
            extra_domains = []
            for row in sheet.iter_rows(values_only=True):
                if row and row[0]:
                    val = str(row[0]).strip().lower()
                    if val and val != "domain" and "." in val:
                        extra_domains.append(val)
            if extra_domains:
                brands.extend(extra_domains)
                brands = list(set(brands))
        except Exception:
            pass

    # Instantiate the analyzers and aggregator
    header_analyzer = HeaderAnalyzer(brand_list=brands)
    url_analyzer = URLAnalyzer(brand_list=brands)
    
    # Load NLP models and ML weights once at startup
    nlp_analyzer = NLPAnalyzer()
    nlp_analyzer.load_models()
    
    aggregator = HybridRiskAggregator()
    aggregator.load_model()
    
    yield

app = FastAPI(
    title="Phishing Filter Microservice",
    description="EPG Member 2 Phishing Filter on port 8002",
    version="1.0.0",
    lifespan=lifespan
)

@app.get("/health")
def health_check():
    return {"status": "healthy", "service": "phishing-filter"}

@app.post("/scan")
async def scan_eml(file: UploadFile = File(...)):
    """
    Accepts a .eml file upload, parses it, runs detection logic, 
    and returns a binary verdict, score, reasons note, and full details.
    """
    # Verify file extension (support basic EML, TXT)
    filename = file.filename.lower() if file.filename else ""
    if not (filename.endswith('.eml') or filename.endswith('.txt') or filename.endswith('.msg') or filename == ""):
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload a standard email (.eml) file.")
    
    temp_file_path = None
    buf, _log_handler = _make_log_capture()
    root_logger = logging.getLogger()
    root_logger.setLevel(logging.INFO)   # force INFO - basicConfig is no-op after uvicorn sets up logging
    root_logger.addHandler(_log_handler)
    try:
        # Create a secure temporary file to write raw upload contents
        with tempfile.NamedTemporaryFile(delete=False, suffix=".eml") as temp_file:
            contents = await file.read()
            temp_file.write(contents)
            temp_file_path = temp_file.name

        # Read contents back as string for EmailParser
        with open(temp_file_path, "r", encoding="utf-8", errors="ignore") as f:
            raw_email_str = f.read()

        logger.info(f"Phishing Scanner: Processing {filename} ({len(contents)} bytes)")
        # Parse EML contents using advanced parser
        email_data = EmailParser.parse_from_string(raw_email_str)

        urls = email_data.get('urls', [])
        logger.info(f"Parsed email — found {len(urls)} URLs")
        if urls:
            for u in urls[:5]:
                logger.info(f"  URL: {u.get('href', '?')}")
        
        # Run pipeline analyzers
        logger.info("[1/3] Running Header Analysis...")
        h_res = header_analyzer.analyze(email_data)
        logger.info(f"  Header score: {h_res['score']}")
        
        logger.info("[2/3] Running URL Analysis...")
        u_res = url_analyzer.analyze(email_data)
        logger.info(f"  URL score: {u_res['score']}")
        
        logger.info("[3/3] Running NLP Analysis...")
        n_res = nlp_analyzer.analyze(email_data)
        logger.info(f"  NLP phishing probability: {n_res['phishing_probability']:.4f}")

        # Run risk aggregator
        logger.info("Aggregating results...")
        agg_res = aggregator.aggregate(h_res, u_res, n_res)
        logger.info(f"Final Phishing Verdict: {agg_res['verdict']} — Risk Score: {agg_res['risk_score']} — Confidence: {agg_res['confidence']}")

        # Structure response strictly according to the format
        verdict = agg_res["verdict"]  # "PHISHING" or "CLEAN"
        score = int(round(agg_res["risk_score"]))  # integer 0-100
        
        # Compile explainable reasons for the note
        reasons_list = agg_res.get("reasons", [])
        note = "; ".join(reasons_list) if reasons_list else "No suspicious indicators detected."

        details = {
            "confidence": agg_res["confidence"],
            "mode": agg_res["mode"],
            "metrics": {
                "header_score": h_res["score"],
                "url_score": u_res["score"],
                "nlp_score": n_res["phishing_probability"] * 100.0
            },
            "header_details": h_res.get("details", {}),
            "urls_detected": [u["href"] for u in email_data.get("urls", [])]
        }

        # --- Generate ASCII Report ---
        report = []
        report.append("")
        report.append("=" * 80)
        report.append("[+] PHISHING ANALYSIS REPORT")
        report.append("=" * 80)
        report.append(f"File: {filename}")
        report.append(f"Verdict: {verdict}")
        report.append(f"Final Score: {score}/100")
        report.append(f"Confidence: {details['confidence']:.1f}%")
        report.append(f"Mode: {details['mode']}")
        report.append("")
        report.append("-" * 28 + " Header Analysis " + "-" * 35)
        report.append(f"Score: {h_res['score']:.1f}/100")
        report.append("")
        report.append("-" * 28 + " URL Analysis " + "-" * 38)
        report.append(f"Score: {u_res['score']:.1f}/100")
        report.append(f"URLs Found: {len(details['urls_detected'])}")
        report.append("")
        report.append("-" * 28 + " NLP Analysis " + "-" * 38)
        report.append(f"Phishing Probability: {details['metrics']['nlp_score']:.2f}%")
        report.append("")
        report.append("Reasons:")
        if reasons_list:
            for r in reasons_list:
                report.append(f"  [!] {r}")
        else:
            report.append("  No suspicious indicators detected.")
        report.append("=" * 80)
        
        for line in report:
            logger.info(line)

        return {
            "verdict": verdict,
            "score": score,
            "note": note,
            "details": details,
            "logs": buf.getvalue()
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Phishing scan failed: {str(e)}")
        
    finally:
        try:
            logging.getLogger().removeHandler(_log_handler)
        except NameError:
            pass
        # Securely clean up the temp file
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except Exception:
                pass

if __name__ == "__main__":
    uvicorn.run("api_server:app", host="0.0.0.0", port=8002)
