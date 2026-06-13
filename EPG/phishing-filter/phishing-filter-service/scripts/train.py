import os
import sys
import random
import csv
from pathlib import Path

# Add project root to path so we can import src
sys.path.append(str(Path(__file__).resolve().parent.parent))

from src.config import DATASET_CSV, DOMAIN_EXCEL, MODELS_DIR, LEGITIMATE_BRANDS
from src.utils.email_parser import EmailParser
from src.utils.pure_ml import PureNaiveBayes, PureLogisticRegression
from src.engines.header_analyzer import HeaderAnalyzer
from src.engines.url_analyzer import URLAnalyzer
from src.engines.nlp_analyzer import NLPAnalyzer
from src.engines.aggregator import HybridRiskAggregator

def preprocess_text(subject, body):
    text = f"{str(subject)} {str(body)}"
    import re
    text = re.sub(r'\s+', ' ', text)
    return text.lower().strip()

def matches_keywords(text, kw_list):
    text_lower = text.lower()
    return any(kw in text_lower for kw in kw_list)

def main():
    print("=== Phishing Pipeline Model Training (Pure Python) ===")
    
    # 1. Check dataset
    if not DATASET_CSV.exists():
        print(f"Error: Dataset not found at {DATASET_CSV}")
        sys.exit(1)
        
    print(f"Loading dataset from {DATASET_CSV} using csv.DictReader (memory safe)...")
    
    rows = []
    try:
        f = open(DATASET_CSV, encoding='utf-8', errors='ignore')
        reader = csv.DictReader(f)
        # Try a quick read to verify
        next(reader)
        f.seek(0)
    except Exception:
        f = open(DATASET_CSV, encoding='latin1', errors='ignore')
        reader = csv.DictReader(f)
        
    for row in reader:
        lbl = str(row.get('label', '')).strip()
        if lbl in ['0', '1']:
            rows.append({
                'sender': row.get('sender') or "",
                'receiver': row.get('receiver') or "",
                'date': row.get('date') or "",
                'subject': row.get('subject') or "",
                'body': row.get('body') or "",
                'label': int(lbl),
                'urls': row.get('urls') or "[]"
            })
    f.close()
    
    print(f"Loaded and parsed {len(rows)} clean rows with binary labels.")
    
    # Calculate class counts
    class_counts = {0: sum(1 for r in rows if r['label'] == 0), 1: sum(1 for r in rows if r['label'] == 1)}
    print(f"Class distribution: {class_counts}")
    
    # Add combined preprocessed text to rows
    for r in rows:
        r['cleaned_text'] = preprocess_text(r['subject'], r['body'])
    
    # 2. Bootstrap Sub-Labels for NLP Multi-Head Classifier
    print("Bootstrapping sub-labels for NLP engines...")
    
    urgency_indicators = [
        "urgent", "immediate", "suspend", "expire", "terminate", "restrict", 
        "action required", "24 hours", "48 hours", "unauthorized", "critical", 
        "final notice", "deadline"
    ]
    
    credential_indicators = [
        "login", "verify", "update", "password", "credential", "sign in", 
        "account details", "confirm", "validation", "reset password", 
        "verify identity", "sign-in"
    ]
    
    bec_indicators = [
        "wire transfer", "payment", "invoice", "gift card", "executive", "ceo", 
        "urgent request", "financial", "direct deposit", "confidential request", 
        "transfer funds"
    ]
    
    for r in rows:
        r['urgency_label'] = 1 if (r['label'] == 1 and matches_keywords(r['cleaned_text'], urgency_indicators)) else 0
        r['credential_label'] = 1 if (r['label'] == 1 and matches_keywords(r['cleaned_text'], credential_indicators)) else 0
        r['bec_label'] = 1 if (r['label'] == 1 and matches_keywords(r['cleaned_text'], bec_indicators)) else 0
        
    print(f"Urgency-labeled rows: {sum(1 for r in rows if r['urgency_label'] == 1)}")
    print(f"Credential-labeled rows: {sum(1 for r in rows if r['credential_label'] == 1)}")
    print(f"BEC-labeled rows: {sum(1 for r in rows if r['bec_label'] == 1)}")
    
    # 3. Train-test Split in Pure Python
    random.seed(42)
    random.shuffle(rows)
    split_idx = int(len(rows) * 0.8)
    train_rows = rows[:split_idx]
    test_rows = rows[split_idx:]
    
    # 4. Train NLP Naive Bayes models
    print("\nTraining Pure Naive Bayes NLP Models...")
    
    # Phishing model
    phishing_clf = PureNaiveBayes()
    phishing_clf.fit([r['cleaned_text'] for r in train_rows], [r['label'] for r in train_rows])
    
    # Evaluate Phishing
    correct = sum(1 for r in test_rows if (phishing_clf.predict_proba(r['cleaned_text']) >= 0.5) == r['label'])
    print(f"Phishing Model Test Accuracy: {correct / len(test_rows):.4f}")
    
    # Urgency model
    urgency_clf = PureNaiveBayes()
    urgency_clf.fit([r['cleaned_text'] for r in train_rows], [r['urgency_label'] for r in train_rows])
    
    # Credential model
    credential_clf = PureNaiveBayes()
    credential_clf.fit([r['cleaned_text'] for r in train_rows], [r['credential_label'] for r in train_rows])
    
    # BEC model
    bec_clf = PureNaiveBayes()
    bec_clf.fit([r['cleaned_text'] for r in train_rows], [r['bec_label'] for r in train_rows])
    
    # Save the NLP models
    print("Saving NLP models to JSON...")
    MODELS_DIR.mkdir(parents=True, exist_ok=True)
    phishing_clf.save(MODELS_DIR / "pure_nlp_phishing.json")
    urgency_clf.save(MODELS_DIR / "pure_nlp_urgency.json")
    credential_clf.save(MODELS_DIR / "pure_nlp_credential.json")
    bec_clf.save(MODELS_DIR / "pure_nlp_bec.json")
    print("NLP Models saved successfully.")
    
    # 5. Extract Brand List from Excel if available
    brands = list(LEGITIMATE_BRANDS)
    if DOMAIN_EXCEL.exists():
        try:
            print(f"Attempting to load legitimate domains from {DOMAIN_EXCEL}...")
            # We can use openpyxl directly
            from openpyxl import load_workbook
            wb = load_workbook(DOMAIN_EXCEL, read_only=True)
            sheet = wb.active
            extra_domains = []
            for row in sheet.iter_rows(values_only=True):
                if row and row[0]:
                    val = str(row[0]).strip().lower()
                    if val and val != "domain" and "." in val:
                        extra_domains.append(val)
            if extra_domains:
                brands.extend(extra_domains)
                brands = list(set(brands))
                print(f"Loaded {len(extra_domains)} brand domains from Excel sheet.")
        except Exception as e:
            print(f"Could not load domain Excel ({e}). Using default brand list.")
            
    # 6. Extract Feature Vectors to Train the Centralized Aggregator
    print("\nExtracting feature vectors for Hybrid Aggregator...")
    header_analyzer = HeaderAnalyzer(brand_list=brands)
    url_analyzer = URLAnalyzer(brand_list=brands)
    
    # Re-initialize NLP analyzer (now it can load the saved Naive Bayes models)
    nlp_analyzer = NLPAnalyzer()
    nlp_analyzer.load_models()
    
    # Initialize a temporary aggregator to get the feature compiler
    aggregator = HybridRiskAggregator()
    
    feature_vectors = []
    labels = []
    
    # Train aggregator on a sample of rows to make training speedy but robust
    sample_size = min(3000, len(rows))
    sample_rows = rows[:sample_size]
    
    count = 0
    for row_data in sample_rows:
        email_data = EmailParser.parse_from_dict(row_data)
        
        # Analyze
        h_res = header_analyzer.analyze(email_data)
        u_res = url_analyzer.analyze(email_data)
        n_res = nlp_analyzer.analyze(email_data)
        
        # Compile vector
        vec = aggregator._compile_feature_vector(h_res, u_res, n_res)
        
        feature_vectors.append(vec)
        labels.append(row_data['label'])
        
        count += 1
        if count % 500 == 0:
            print(f"Processed {count}/{sample_size} rows for aggregator features.")
            
    # Split aggregator dataset into train/test
    split_idx_agg = int(len(feature_vectors) * 0.8)
    X_train_agg = feature_vectors[:split_idx_agg]
    y_train_agg = labels[:split_idx_agg]
    X_test_agg = feature_vectors[split_idx_agg:]
    y_test_agg = labels[split_idx_agg:]
    
    print("Training Pure Logistic Regression Aggregator...")
    agg_model = PureLogisticRegression(lr=0.05, iters=800)
    agg_model.fit(X_train_agg, y_train_agg)
    
    # Evaluate Aggregator
    correct_agg = 0
    tp, fp, fn, tn = 0, 0, 0, 0
    for x, y in zip(X_test_agg, y_test_agg):
        pred_prob = agg_model.predict_proba(x)
        pred_label = 1 if pred_prob >= 0.5 else 0
        if pred_label == y:
            correct_agg += 1
        
        if pred_label == 1 and y == 1:
            tp += 1
        elif pred_label == 1 and y == 0:
            fp += 1
        elif pred_label == 0 and y == 1:
            fn += 1
        elif pred_label == 0 and y == 0:
            tn += 1
            
    accuracy = correct_agg / len(X_test_agg)
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
    
    print("\n--- Aggregator Model Evaluation ---")
    print(f"Accuracy:  {accuracy:.4f}")
    print(f"Precision: {precision:.4f}")
    print(f"Recall:    {recall:.4f}")
    print(f"F1-Score:  {f1:.4f}")
    
    # Save the Aggregator model
    agg_model.save(MODELS_DIR / "pure_aggregator.json")
    print("Aggregator Model saved successfully.")
    print("=== Training completed successfully! ===")

if __name__ == "__main__":
    main()
