import sys
import argparse
import json
from pathlib import Path

# Add project root to python path
sys.path.append(str(Path(__file__).resolve().parent))

from src.utils.email_parser import EmailParser
from src.engines.header_analyzer import HeaderAnalyzer
from src.engines.url_analyzer import URLAnalyzer
from src.engines.nlp_analyzer import NLPAnalyzer
from src.engines.aggregator import HybridRiskAggregator
from src.config import LEGITIMATE_BRANDS, DOMAIN_EXCEL

def main():
    parser = argparse.ArgumentParser(description="Advanced Phishing Detection Pipeline CLI")
    parser.add_argument("--file", "-f", type=str, help="Path to raw .eml file to analyze")
    parser.add_argument("--json-out", "-j", action="store_true", help="Output results in JSON format")
    args = parser.parse_args()

    # Reconfigure stdout to support unicode prints on Windows
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')

    # ASCII Art Banner
    if not args.json_out:
        print("""
====================================================================
[SHIELD] ADVANCED PHISHING DETECTION PIPELINE - SOC CLI RUNNER
====================================================================
        """)

    brands = list(LEGITIMATE_BRANDS)
    if DOMAIN_EXCEL.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(DOMAIN_EXCEL, read_only=True)
            sheet = wb.active
            extra_domains = []
            for row in sheet.iter_rows(values_only=True):
                if row and row[0]:
                    val = str(row[0]).strip().lower()
                    if val and val != "domain" and "." in val:
                        extra_domains.append(val)
            if extra_domains:
                brands.extend(extra_domains)
                brands = list(set(brands))
        except Exception:
            pass

    header_analyzer = HeaderAnalyzer(brand_list=brands)
    url_analyzer = URLAnalyzer(brand_list=brands)
    nlp_analyzer = NLPAnalyzer()
    aggregator = HybridRiskAggregator()

    email_data = None
    if args.file:
        file_path = Path(args.file)
        if not file_path.exists():
            print(f"[-] Error: File not found at {args.file}")
            sys.exit(1)
        
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                raw_email_str = f.read()
            email_data = EmailParser.parse_from_string(raw_email_str)
        except Exception as e:
            print(f"[-] Error reading email file: {e}")
            sys.exit(1)
    else:
        # If no file provided, run on a default suspicious email snippet for demonstration
        if not args.json_out:
            print("[*] No input file specified. Running demonstration scan on mock phishing email...")
        email_data = {
            "sender": '"PayPal Security Services" <security-compliance@paypal-resolution-center.net>',
            "receiver": "user@target.com",
            "date": "Wed, 20 May 2026 21:00:00 +0300",
            "subject": "Action Required: Unusual Activity Detected on your PayPal Account",
            "reply_to": "hacker-box@gmail.com",
            "x_mailer": "PHP mailer script",
            "body_text": "Your account has been suspended due to policy violations. Please verify your account immediately at http://paypal-resolution.com/login.html",
            "body_html": 'Your account has been suspended. Please <a href="http://paypal-resolution.com/login.html">verify your account</a> immediately.',
            "urls": [
                {"href": "http://paypal-resolution.com/login.html", "text": "verify your account"}
            ],
            "auth_headers": {
                "spf": "FAIL",
                "dkim": "NONE",
                "dmarc": "FAIL",
                "raw_auth_results": "spf=fail; dmarc=fail"
            }
        }

    # Run analysis
    h_res = header_analyzer.analyze(email_data)
    u_res = url_analyzer.analyze(email_data)
    n_res = nlp_analyzer.analyze(email_data)
    agg_res = aggregator.aggregate(h_res, u_res, n_res)

    if args.json_out:
        # Print pure json response for integrations
        output = {
            "verdict": agg_res["verdict"],
            "risk_score": agg_res["risk_score"],
            "confidence": agg_res["confidence"],
            "reasons": agg_res["reasons"],
            "mode": agg_res["mode"],
            "breakdown": {
                "header_score": h_res["score"],
                "url_score": u_res["score"],
                "nlp_score": n_res["phishing_probability"] * 100.0
            }
        }
        print(json.dumps(output, indent=2))
    else:
        # Display formatted text report
        verdict = agg_res["verdict"]
        score = agg_res["risk_score"]
        conf = agg_res["confidence"]
        
        # Color formatting — binary only: PHISHING (red) or CLEAN (green)
        verdict_str = f"\033[91m{verdict}\033[0m" if verdict == "PHISHING" else f"\033[92m{verdict}\033[0m"
        
        print(f"VERDICT:      {verdict_str}")
        print(f"RISK SCORE:   {score:.1f} / 100.0")
        print(f"CONFIDENCE:   {conf * 100:.0f}%")
        print(f"ENGINE MODE:  {agg_res['mode']}")
        print("\n--- Sub-Engine Breakdowns ---")
        print(f"Header Integrity Score: {h_res['score']:.1f}/100")
        print(f"URL Obfuscation Score:  {u_res['score']:.1f}/100")
        print(f"NLP Semantic Score:     {n_res['phishing_probability'] * 100:.1f}/100")
        
        print("\n--- Explainable AI Reasons ---")
        if agg_res["reasons"]:
            for r in agg_res["reasons"]:
                print(f" - [!] {r}")
        else:
            print(" - [PASS] No policy flags triggered.")
            
        print("\n--- Header Information ---")
        print(f"Display Name:   {h_res['details']['display_name']}")
        print(f"Sender Address: {h_res['details']['sender_address']}")
        print(f"SPF Status:     {h_res['details']['auth_headers'].get('spf')}")
        print(f"DMARC Status:   {h_res['details']['auth_headers'].get('dmarc')}")
        print(f"X-Mailer:       {h_res['details']['x_mailer']}")

        print("\n====================================================================")

if __name__ == "__main__":
    main()
